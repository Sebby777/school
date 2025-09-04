import pandas as pd
import numpy as np
import os
from openpyxl import load_workbook
import shutil
import re

def preprocess_student_data(student_file_path):
    """
    预处理学生数据表，找到所需的列
    """
    # 先读取前几行，查找表头行
    preview_df = pd.read_excel(student_file_path, header=None)
    school_keywords = ['学校', '幼儿园', '小学', '中学', '高中', '统筹点']
    school_name = None
    header_row = None

    # 查找学校名称
    for i in range(min(5, len(preview_df))):
        row = preview_df.iloc[i]
        for cell in row:
            if isinstance(cell, str):
                for kw in school_keywords:
                    if kw in cell:
                        # 用正则只提取“xxx学校”或“xxx幼儿园”等
                        match = re.search(r'([^\s：]+(学校|幼儿园|小学|中学|高中))', cell)
                        if match:
                            school_name = match.group(1)
                        else:
                            # 如果没有匹配到，去掉前缀“学校名称/统筹点：”
                            school_name = re.sub(r'^.*?：', '', cell).strip()
                        break
            if school_name:
                break
        if school_name:
            break
    # 查找表头行
    for i in range(min(10, len(preview_df))):
        row = preview_df.iloc[i]
        if ('班级' in row.values) and ('姓名' in row.values):
            header_row = i
            break
    if header_row is None:
        raise ValueError("未找到表头行，请检查Excel格式")

    # 用 header_row 重新读取数据
    df = pd.read_excel(student_file_path, header=header_row)

    # 补充关键词，支持括号和单位
    keyword_columns = {
        '班级': ['班级', '班别'],
        '姓名': ['姓名', '名字'],
        '性别': ['性别'],
        '年龄': ['年龄'],
        '身高': ['身高', '身高(cm)', '身高（cm）'],
        '体重': ['体重', '体重(kg)', '体重（kg）', '体重(斤)', '体重（斤）'],
        '备注': ['备注', '说明']
    }

    # 尝试找到包含这些关键词的列
    col_mapping = {}
    for std_col, possible_names in keyword_columns.items():
        for col in df.columns:
            if any(name in str(col) for name in possible_names):
                col_mapping[std_col] = col
                break

    necessary_cols = ['班级', '姓名', '性别', '年龄', '身高', '体重', '备注']
    if all(col in col_mapping for col in necessary_cols):
        processed_df = df[list(col_mapping.values())].copy()
        processed_df.columns = necessary_cols
        processed_df['学校'] = school_name if school_name else ''
        
        # 判断体重列名是否有“斤”
        weight_col_name = col_mapping['体重']
        is_jin = '斤' in str(weight_col_name)
        
        def convert_weight(val):
            if pd.isna(val):
                return np.nan
            num = re.findall(r'\d+\.?\d*', str(val))
            if not num:
                return np.nan
            if is_jin:
                return float(num[0]) / 2
            else:
                return float(num[0])
        
        processed_df['体重'] = processed_df['体重'].apply(convert_weight)
        return processed_df
    else:
        print("实际表头：", df.columns.tolist())
        print("已识别列：", col_mapping)
        raise ValueError("无法识别学生数据表的列结构")

def load_size_table(file_path):
    """
    加载码数表并解析为结构化数据
    """
    import pandas as pd
    import re
    
    # 读取码数表
    df = pd.read_excel(file_path, header=None)
    
    # 解析儿童打底码数
    children_underwear = []
    start_row = None
    for i in range(len(df)):
        if '儿童打底码数尺寸' in str(df.iloc[i, 0]):
            start_row = i + 2  # 跳过标题行
            break
    
    if start_row is not None:
        for i in range(start_row, len(df)):
            if pd.isna(df.iloc[i, 0]) or '成人' in str(df.iloc[i, 0]):
                break
            size = df.iloc[i, 0]
            weight_range = df.iloc[i, 1]
            height_range = df.iloc[i, 2]
            
            # 检查是否为空值
            if pd.isna(size) or pd.isna(weight_range) or pd.isna(height_range):
                continue
                
            # 解析体重范围
            weight_range = str(weight_range).replace('斤', '')
            if '-' not in weight_range:
                continue
            try:
                weight_min, weight_max = map(float, weight_range.split('-'))
            except Exception:
                continue
            
            # 解析身高范围
            height_range = str(height_range).replace('cm', '')
            if '-' not in height_range:
                continue
            try:
                height_min, height_max = map(float, height_range.split('-'))
            except Exception:
                continue
            
            children_underwear.append({
                'type': '儿童',
                'category': '打底',
                'size': size,
                'weight_min': weight_min,
                'weight_max': weight_max,
                'height_min': height_min,
                'height_max': height_max
            })
    
    # 解析儿童棉衣码数
    children_coat = []
    start_row = None
    for i in range(len(df)):
        if '儿童棉衣码数尺寸' in str(df.iloc[i, 3]):
            start_row = i + 2  # 跳过标题行
            break
    
    if start_row is not None:
        for i in range(start_row, len(df)):
            if pd.isna(df.iloc[i, 3]) or '成人' in str(df.iloc[i, 3]):
                break
                
            size = df.iloc[i, 3]
            weight_range = df.iloc[i, 4]
            age_range = df.iloc[i, 5]
            
            # 检查是否为空值
            if pd.isna(size) or pd.isna(weight_range) or pd.isna(age_range):
                continue
                
            # 解析体重范围
            weight_range = str(weight_range).replace('斤', '')
            if '-' not in weight_range:
                continue
            try:
                weight_min, weight_max = map(float, weight_range.split('-'))
            except Exception:
                continue
            
            # 解析年龄范围
            age_range = str(age_range).replace('岁', '')
            if '-' not in age_range:
                continue
            try:
                age_min, age_max = map(float, age_range.split('-'))
            except Exception:
                continue
            
            children_coat.append({
                'type': '儿童',
                'category': '棉衣',
                'size': size,
                'weight_min': weight_min,
                'weight_max': weight_max,
                'age_min': age_min,
                'age_max': age_max
            })
    
    # 解析成人打底码数
    adult_underwear = []
    start_row = None
    for i in range(len(df)):
        if '成人打底码数' in str(df.iloc[i, 0]):
            start_row = i + 1  # 跳过标题行
            break
    
    if start_row is not None:
        gender = None
        for i in range(start_row, len(df)):
            cell_value = str(df.iloc[i, 0])
            
            if '男生' in cell_value:
                gender = '男'
                # 男生第一行是L码
                if not pd.isna(df.iloc[i, 1]):
                    weight_range = df.iloc[i, 1]
                    weight_range = str(weight_range).replace('斤', '')
                    if '-' in weight_range:
                        try:
                            weight_min, weight_max = map(float, weight_range.split('-'))
                            adult_underwear.append({
                                'type': '成人',
                                'category': '打底',
                                'gender': gender,
                                'size': 'L',
                                'weight_min': weight_min,
                                'weight_max': weight_max
                            })
                        except Exception:
                            pass
                continue
            elif '女生' in cell_value:
                gender = '女'
                continue
            elif pd.isna(df.iloc[i, 0]) or not cell_value.strip():
                continue
            elif '成人棉衣尺寸' in cell_value:
                break
            
            size = df.iloc[i, 0]
            weight_range = df.iloc[i, 1]
            
            # 检查是否为空值
            if pd.isna(weight_range):
                continue
                
            # 解析体重范围
            weight_range = str(weight_range).replace('斤', '')
            if '-' not in weight_range:
                continue
            try:
                weight_min, weight_max = map(float, weight_range.split('-'))
            except Exception:
                continue
            
            adult_underwear.append({
                'type': '成人',
                'category': '打底',
                'gender': gender,
                'size': size,
                'weight_min': weight_min,
                'weight_max': weight_max
            })
    
    # 解析成人棉衣码数
    adult_coat = []
    start_row = None
    for i in range(len(df)):
        if '成人棉衣尺寸' in str(df.iloc[i, 3]):
            start_row = i + 1  # 跳过标题行
            break
    
    if start_row is not None:
        for i in range(start_row, len(df)):
            if pd.isna(df.iloc[i, 3]) or not str(df.iloc[i, 3]).strip():
                break
                
            size = df.iloc[i, 3]
            # 确保尺码是大写
            if isinstance(size, str):
                size = size.upper()
                
            weight_range = df.iloc[i, 4]
            
            # 检查是否为空值
            if pd.isna(weight_range):
                continue
                
            # 解析体重范围
            weight_range = str(weight_range).replace('斤', '')
            if '-' not in weight_range:
                continue
            try:
                weight_min, weight_max = map(float, weight_range.split('-'))
            except Exception:
                continue
            
            adult_coat.append({
                'type': '成人',
                'category': '棉衣',
                'size': size,
                'weight_min': weight_min,
                'weight_max': weight_max
            })
    
    # 解析袜子尺码
    socks_sizes = []
    start_row = None
    for i in range(len(df)):
        if '袜子尺码' in str(df.iloc[i, 6]):
            start_row = i + 1  # 跳过标题行
            break
    
    if start_row is not None:
        for i in range(start_row, len(df)):
            if pd.isna(df.iloc[i, 6]) or not str(df.iloc[i, 6]).strip():
                continue
            
            age_range = df.iloc[i, 6]
            size_label = age_range  # 使用年龄范围作为尺码标签
            
            # 处理特殊情况：8个月-2岁
            if '8个月-2岁' in str(age_range):
                socks_sizes.append({
                    'category': '袜子',
                    'size': size_label,
                    'age_min': 0.67,  # 8个月 ≈ 0.67岁
                    'age_max': 2.0
                })
                continue
                
            # 处理其他年龄范围
            age_range = str(age_range).replace('岁', '').strip()
            if '-' in age_range:
                try:
                    age_parts = age_range.split('-')
                    age_min = float(age_parts[0])
                    age_max = float(age_parts[1])
                    socks_sizes.append({
                        'category': '袜子',
                        'size': size_label,
                        'age_min': age_min,
                        'age_max': age_max
                    })
                except Exception:
                    continue
    
    return {
        'children_underwear': children_underwear,
        'children_coat': children_coat,
        'adult_underwear': adult_underwear,
        'adult_coat': adult_coat,
        'socks_sizes': socks_sizes
    }

def recommend_size(student, size_data, category):
    """
    根据学生信息和码数表推荐合适的码数
    category: 'coat' 或 'underwear'
    只考虑身高和体重，不考虑年龄
    """
    height = student['身高']
    weight = student['体重'] * 2  # 转换为斤
    gender = student['性别']

    # 首先尝试儿童码数
    if category == 'coat':
        size_list = size_data['children_coat']
        adult_list = size_data['adult_coat']
    else:
        size_list = size_data['children_underwear']
        adult_list = size_data['adult_underwear']

    # 找出所有合适的儿童码数（只考虑身高和体重）
    suitable_sizes = []
    for item in size_list:
        # 检查体重是否在范围内
        weight_ok = item['weight_min'] <= weight <= item['weight_max']
        
        # 检查身高是否在范围内（如果有身高信息）
        if 'height_min' in item and 'height_max' in item:
            height_ok = item['height_min'] <= height <= item['height_max']
            # 身高或体重有一项满足就考虑
            if weight_ok or height_ok:
                suitable_sizes.append(item)
        elif weight_ok:
            suitable_sizes.append(item)

    # 如果有合适的儿童码数，选择最大的（因为儿童会长大）
    if suitable_sizes:
        # 按码数大小排序（数值或字母）
        def get_size_value(size):
            if isinstance(size, int):
                return size
            elif isinstance(size, str) and size.isdigit():
                return int(size)
            elif isinstance(size, str):
                # 处理字母码数如L, XL等
                size_map = {'S': 1, 'M': 2, 'L': 3, 'XL': 4, 'XXL': 5, '3XL': 6, '4XL': 7}
                return size_map.get(size, 0)
            return 0
        
        # 选择最大的码数（考虑到儿童会长大）
        suitable_sizes.sort(key=lambda x: get_size_value(x['size']), reverse=True)
        return suitable_sizes[0]['size']

    # 如果没有找到合适的儿童码数，使用成人码数
    suitable_adult = []
    for item in adult_list:
        if item['weight_min'] <= weight <= item['weight_max']:
            if 'gender' in item and item['gender'] != gender:
                continue
            suitable_adult.append(item)
    
    if suitable_adult:
        # 按体重上限排序，选择最小合适码
        suitable_adult.sort(key=lambda x: x['weight_max'])
        return suitable_adult[0]['size']
    else:
        # 没有合适区间，返回最接近的码
        adult_list_sorted = sorted(adult_list, key=lambda x: x['weight_min'])
        if weight < adult_list_sorted[0]['weight_min']:
            return adult_list_sorted[0]['size']
        else:
            return adult_list_sorted[-1]['size']
def recommend_socks_size(student, socks_sizes):
    """
    推荐袜子尺码
    """
    age = student['年龄']
    for item in socks_sizes:
        if item['age_min'] <= age <= item['age_max']:
            return item['size']
    if socks_sizes:
        # 如果年龄比所有区间都大，返回最大区间；比所有区间都小，返回最小区间
        socks_sizes.sort(key=lambda x: x['age_min'])
        if age < socks_sizes[0]['age_min']:
            return socks_sizes[0]['size']
        else:
            return socks_sizes[-1]['size']
    return '6-8岁'

def create_result_template():
    """ 
    创建结果模板
    """
    # 创建一个新的DataFrame作为模板
    template_data = {
        '序号': [], '班级': [], '姓名': [], '性别': [], '年龄': [], 
        '身高(cm)': [], '体重(kg)': [], '': [], '学校名称 （统筹点，必填）': [],
        '棉衣配码4': [], '打底配码2': [], '袜子配码': [], '备注': []
    }
    
    return pd.DataFrame(template_data)
def write_result_to_template(result_df, template_path, sheet_name='衣服配码'):
    wb = load_workbook(template_path)
    ws = wb[sheet_name]

    # 假设表头在第3行，数据从第4行开始
    start_row = 4

    columns = [
        '序号', '班级', '姓名', '性别', '年龄', '身高(cm)', '体重(kg)', '', 
        '学校名称 （统筹点，必填）', '棉衣配码4', '打底配码2', '袜子配码', '备注'
    ]
    for i, row in result_df.iterrows():
        for j, col in enumerate(columns, 1):
            ws.cell(row=start_row + i, column=j, value=row[col])

    wb.save(template_path)
    print(f"已追加写入结果到模板: {template_path}")

def main(student_file_path, size_file_path, template_file_path, output_dir):
    # 读取学生数据
    student_df = preprocess_student_data(student_file_path)
    # 加载码数表
    size_data = load_size_table(size_file_path)
    # 创建结果模板
    result_df = create_result_template()
    # 推荐码数
    for idx, student in student_df.iterrows():
        coat_size = recommend_size(student, size_data, 'coat')
        underwear_size = recommend_size(student, size_data, 'underwear')
        socks_size = recommend_socks_size(student, size_data['socks_sizes'])
        new_row = {
            '序号': idx + 1,
            '班级': student['班级'],
            '姓名': student['姓名'],
            '性别': student['性别'],
            '年龄': student['年龄'],
            '身高(cm)': student['身高'],
            '体重(kg)': student['体重'],
            '': student['体重'] * 2,
            '学校名称 （统筹点，必填）': student['学校'],
            '棉衣配码4': coat_size,
            '打底配码2': underwear_size,
            '袜子配码': socks_size,
            '备注': student['备注'] if pd.notna(student['备注']) else ''
        }
        result_df = pd.concat([result_df, pd.DataFrame([new_row])], ignore_index=True)

    # 复制模板文件，命名为学校名，并保存到用户指定的输出目录
    school_name = student_df.iloc[0]['学校']
    match = re.search(r'([^\s：]+学校)', school_name)
    if match:
        pure_school_name = match.group(1)
    else:
        pure_school_name = school_name
    safe_school_name = re.sub(r'[\/\\\:\*\?\"\<\>\|]', '_', pure_school_name)
    output_file_path = os.path.join(output_dir, f"{safe_school_name}.xlsx")
    shutil.copy(template_file_path, output_file_path)

    # 写入结果
    write_result_to_template(result_df, output_file_path)



